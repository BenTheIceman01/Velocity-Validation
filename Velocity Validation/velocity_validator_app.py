"""
HD Supply™ Velocity Validator
Developed by: Ben F. Benjamaa

A modern application for validating velocity codes against Snowflake data.
"""

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
import snowflake.connector
import os
from datetime import datetime
import threading
from tkinter import font as tkfont

class ModernButton(tk.Canvas):
    """Custom modern button widget with hover effects"""
    def __init__(self, parent, text, command, bg_color, fg_color, hover_color, **kwargs):
        width = kwargs.pop('width', 200)
        height = kwargs.pop('height', 50)
        super().__init__(parent, width=width, height=height, bg=parent['bg'], 
                        highlightthickness=0, **kwargs)
        
        self.command = command
        self.bg_color = bg_color
        self.fg_color = fg_color
        self.hover_color = hover_color
        self.text = text
        
        # Create button rectangle
        self.rect = self.create_rectangle(0, 0, width, height, 
                                          fill=bg_color, outline="")
        self.text_item = self.create_text(width/2, height/2, 
                                         text=text, fill=fg_color,
                                         font=("Segoe UI", 11, "bold"))
        
        # Bind events
        self.bind("<Button-1>", lambda e: self.command())
        self.bind("<Enter>", self.on_enter)
        self.bind("<Leave>", self.on_leave)
        self.configure(cursor="hand2")
        
    def on_enter(self, e):
        self.itemconfig(self.rect, fill=self.hover_color)
        
    def on_leave(self, e):
        self.itemconfig(self.rect, fill=self.bg_color)

class VelocityValidatorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("HD Supply™ Velocity Validator")
        self.root.geometry("800x700")
        self.root.resizable(False, False)
        
        # Modern HD Supply color scheme - Black background with Yellow accents
        self.bg_black = "#000000"
        self.hd_yellow = "#FFD700"
        self.hd_bright_yellow = "#FFED4E"
        self.dark_gray = "#1A1A1A"
        self.medium_gray = "#2D2D2D"
        self.light_gray = "#404040"
        self.text_gray = "#CCCCCC"
        
        # Configure root window
        self.root.configure(bg=self.bg_black)
        
        # Variables
        self.input_file_path = tk.StringVar()
        self.snowflake_data = None
        
        # Configure custom styles
        self.setup_styles()
        self.setup_gui()
        
    def setup_styles(self):
        """Setup custom ttk styles"""
        style = ttk.Style()
        style.theme_use('clam')
        
        # Configure Progressbar
        style.configure("Yellow.Horizontal.TProgressbar",
                       troughcolor=self.medium_gray,
                       background=self.hd_yellow,
                       bordercolor=self.bg_black,
                       lightcolor=self.hd_yellow,
                       darkcolor=self.hd_yellow)
        
    def setup_gui(self):
        # Main container with padding
        main_container = tk.Frame(self.root, bg=self.bg_black)
        main_container.pack(fill="both", expand=True, padx=0, pady=0)
        
        # Header section with company branding
        header_frame = tk.Frame(main_container, bg=self.bg_black, height=120)
        header_frame.pack(fill="x", padx=0, pady=0)
        header_frame.pack_propagate(False)
        
        # Company name with TM
        company_frame = tk.Frame(header_frame, bg=self.bg_black)
        company_frame.pack(pady=(20, 5))
        
        company_label = tk.Label(
            company_frame,
            text="HD SUPPLY",
            font=("Segoe UI", 32, "bold"),
            bg=self.bg_black,
            fg=self.hd_bright_yellow
        )
        company_label.pack(side="left")
        
        tm_label = tk.Label(
            company_frame,
            text="™",
            font=("Segoe UI", 16, "bold"),
            bg=self.bg_black,
            fg=self.hd_bright_yellow
        )
        tm_label.pack(side="left", anchor="n", padx=(2, 0))
        
        # Title
        title_label = tk.Label(
            header_frame,
            text="VELOCITY VALIDATOR",
            font=("Segoe UI", 18, "bold"),
            bg=self.bg_black,
            fg=self.hd_yellow
        )
        title_label.pack(pady=(0, 5))
        
        # Developer credit
        dev_label = tk.Label(
            header_frame,
            text="Developed by: Ben F. Benjamaa",
            font=("Segoe UI", 9, "italic"),
            bg=self.bg_black,
            fg=self.text_gray
        )
        dev_label.pack()
        
        # Separator line
        separator = tk.Frame(main_container, bg=self.hd_yellow, height=2)
        separator.pack(fill="x", padx=20, pady=(0, 20))
        
        # Content frame with padding
        content_frame = tk.Frame(main_container, bg=self.bg_black)
        content_frame.pack(fill="both", expand=True, padx=40, pady=0)
        
        # File selection section
        self.create_section(content_frame, "STEP 1: SELECT FILE", 0)
        
        file_info_frame = tk.Frame(content_frame, bg=self.dark_gray)
        file_info_frame.pack(fill="x", pady=(0, 20))
        
        self.file_label = tk.Label(
            file_info_frame,
            text="No file selected",
            bg=self.dark_gray,
            fg=self.text_gray,
            font=("Segoe UI", 10),
            anchor="w",
            padx=20,
            pady=15
        )
        self.file_label.pack(fill="x")
        
        browse_btn = ModernButton(
            content_frame,
            text="BROWSE FILE",
            command=self.browse_file,
            bg_color=self.hd_yellow,
            fg_color=self.bg_black,
            hover_color=self.hd_bright_yellow,
            width=200,
            height=45
        )
        browse_btn.pack(pady=(10, 20))
        
        # Snowflake connection section
        self.create_section(content_frame, "STEP 2: SNOWFLAKE CONNECTION", 20)
        
        sf_frame = tk.Frame(content_frame, bg=self.dark_gray)
        sf_frame.pack(fill="x", pady=(0, 20), padx=0)
        
        # Create input fields
        fields = [
            ("Account", "account", False),
            ("Username", "username", False),
            ("Password", "password", True),
            ("Warehouse", "warehouse", False),
            ("Database", "database", False),
            ("Schema", "schema", False)
        ]
        
        self.sf_inputs = {}
        for i, (label_text, field, is_password) in enumerate(fields):
            field_frame = tk.Frame(sf_frame, bg=self.dark_gray)
            field_frame.pack(fill="x", padx=20, pady=8)
            
            label = tk.Label(
                field_frame,
                text=label_text + ":",
                bg=self.dark_gray,
                fg=self.hd_yellow,
                font=("Segoe UI", 10, "bold"),
                width=12,
                anchor="w"
            )
            label.pack(side="left", padx=(0, 10))
            
            entry = tk.Entry(
                field_frame,
                bg=self.medium_gray,
                fg=self.hd_bright_yellow,
                font=("Segoe UI", 10),
                insertbackground=self.hd_yellow,
                relief="flat",
                bd=0,
                show="●" if is_password else ""
            )
            entry.pack(side="left", fill="x", expand=True, ipady=8, ipadx=10)
            self.sf_inputs[field] = entry
            
            # Set default values
            if field == "database":
                entry.insert(0, "EDP")
            elif field == "schema":
                entry.insert(0, "STD_JDA")
        
        # Process button
        process_btn = ModernButton(
            content_frame,
            text="PROCESS DATA",
            command=self.process_data,
            bg_color=self.hd_yellow,
            fg_color=self.bg_black,
            hover_color=self.hd_bright_yellow,
            width=250,
            height=55
        )
        process_btn.pack(pady=20)
        self.process_btn = process_btn
        
        # Progress bar
        self.progress = ttk.Progressbar(
            content_frame,
            mode="indeterminate",
            length=400,
            style="Yellow.Horizontal.TProgressbar"
        )
        
        # Status label
        self.status_label = tk.Label(
            content_frame,
            text="",
            bg=self.bg_black,
            fg=self.hd_yellow,
            font=("Segoe UI", 10, "bold")
        )
        self.status_label.pack(pady=10)
        
        # Footer
        footer_frame = tk.Frame(main_container, bg=self.bg_black, height=40)
        footer_frame.pack(fill="x", side="bottom")
        footer_frame.pack_propagate(False)
        
        footer_separator = tk.Frame(footer_frame, bg=self.hd_yellow, height=1)
        footer_separator.pack(fill="x", padx=20)
        
        footer_label = tk.Label(
            footer_frame,
            text=f"© {datetime.now().year} HD Supply™ | Version 1.0",
            font=("Segoe UI", 8),
            bg=self.bg_black,
            fg=self.text_gray
        )
        footer_label.pack(pady=10)
        
    def create_section(self, parent, title, pady_top):
        """Create a section header"""
        section_label = tk.Label(
            parent,
            text=title,
            font=("Segoe UI", 11, "bold"),
            bg=self.bg_black,
            fg=self.hd_bright_yellow,
            anchor="w"
        )
        section_label.pack(fill="x", pady=(pady_top, 10))
        
    def browse_file(self):
        filename = filedialog.askopenfilename(
            title="Select Excel/CSV file",
            filetypes=[
                ("CSV files", "*.csv"),
                ("Excel files", "*.xlsx *.xls"),
                ("All files", "*.*")
            ]
        )
        if filename:
            self.input_file_path.set(filename)
            display_name = os.path.basename(filename)
            if len(display_name) > 50:
                display_name = display_name[:47] + "..."
            self.file_label.config(text=f"✓ {display_name}")
            
    def connect_snowflake(self):
        """Connect to Snowflake and fetch velocity data"""
        try:
            conn = snowflake.connector.connect(
                account=self.sf_inputs['account'].get().strip(),
                user=self.sf_inputs['username'].get().strip(),
                password=self.sf_inputs['password'].get(),
                warehouse=self.sf_inputs['warehouse'].get().strip(),
                database=self.sf_inputs['database'].get().strip(),
                schema=self.sf_inputs['schema'].get().strip()
            )
            
            query = """
            SELECT
                ITEM,
                LOC,
                UDC_VELOCITY_CODE
            FROM
                SKUEXTRACT
            """
            
            self.snowflake_data = pd.read_sql(query, conn)
            conn.close()
            return True
            
        except Exception as e:
            self.root.after(0, lambda: messagebox.showerror(
                "Connection Error", 
                f"Failed to connect to Snowflake:\n\n{str(e)}"
            ))
            return False
            
    def validate_inputs(self):
        """Validate all required inputs"""
        if not self.input_file_path.get():
            messagebox.showwarning("Missing Input", "Please select an Excel/CSV file!")
            return False
            
        required_fields = ['account', 'username', 'password', 'warehouse', 'database', 'schema']
        for field in required_fields:
            if not self.sf_inputs[field].get().strip():
                messagebox.showwarning("Missing Input", f"Please enter {field.replace('_', ' ').title()}!")
                return False
                
        return True
            
    def process_data(self):
        """Start data processing in a separate thread"""
        if not self.validate_inputs():
            return
            
        # Run processing in a separate thread to keep UI responsive
        thread = threading.Thread(target=self.process_data_thread, daemon=True)
        thread.start()
        
    def process_data_thread(self):
        """Process the data in a background thread"""
        try:
            # Update UI from main thread
            self.root.after(0, self.start_processing_ui)
            
            # Connect to Snowflake
            self.root.after(0, lambda: self.status_label.config(
                text="⚡ Connecting to Snowflake..."
            ))
            
            if not self.connect_snowflake():
                self.root.after(0, self.stop_processing_ui)
                return
                
            self.root.after(0, lambda: self.status_label.config(
                text="⚡ Loading data file..."
            ))
            
            # Read the input file
            file_path = self.input_file_path.get()
            if file_path.endswith('.csv'):
                df = pd.read_csv(file_path)
            else:
                df = pd.read_excel(file_path)
                
            self.root.after(0, lambda: self.status_label.config(
                text="⚡ Processing and validating data..."
            ))
            
            # Perform VLOOKUP - merge on ITEM and LOC
            if 'ITEM' not in df.columns or 'LOC' not in df.columns:
                self.root.after(0, lambda: messagebox.showerror(
                    "Column Error",
                    "Required columns ITEM and/or LOC not found in input file!"
                ))
                self.root.after(0, self.stop_processing_ui)
                return
                
            # Merge data
            df_merged = df.merge(
                self.snowflake_data,
                on=['ITEM', 'LOC'],
                how='left'
            )
            
            # Rename the velocity code column
            df_merged.rename(columns={'UDC_VELOCITY_CODE': 'Current_Velocity'}, inplace=True)
            
            # Add Match column (True/False comparison)
            if 'PROPOSED_VELOCITY' in df_merged.columns:
                df_merged['Match'] = df_merged.apply(
                    lambda row: row['Current_Velocity'] == row['PROPOSED_VELOCITY'] 
                    if pd.notna(row['Current_Velocity']) and pd.notna(row['PROPOSED_VELOCITY'])
                    else False,
                    axis=1
                )
            else:
                self.root.after(0, lambda: messagebox.showwarning(
                    "Warning",
                    "PROPOSED_VELOCITY column not found in input file.\nMatch column will be set to False."
                ))
                df_merged['Match'] = False
                
            self.root.after(0, lambda: self.status_label.config(
                text="⚡ Generating Excel report..."
            ))
            
            # Generate output filename
            input_dir = os.path.dirname(file_path)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"Velocity_Validated_{timestamp}.xlsx"
            output_path = os.path.join(input_dir, output_filename)
            
            # Save to Excel with formatting
            self.save_formatted_excel(df_merged, output_path)
            
            # Calculate statistics
            total_rows = len(df_merged)
            matches = df_merged['Match'].sum() if 'Match' in df_merged.columns else 0
            mismatches = total_rows - matches
            
            self.root.after(0, self.stop_processing_ui)
            self.root.after(0, lambda: self.status_label.config(
                text=f"✓ SUCCESS! File saved: {output_filename}"
            ))
            
            self.root.after(0, lambda: messagebox.showinfo(
                "Processing Complete",
                f"✓ Velocity validation completed successfully!\n\n"
                f"Output: {output_filename}\n\n"
                f"Statistics:\n"
                f"• Total Records: {total_rows:,}\n"
                f"• Matches: {matches:,}\n"
                f"• Mismatches: {mismatches:,}\n\n"
                f"Columns Added:\n"
                f"• Current_Velocity (from Snowflake)\n"
                f"• Match (True/False comparison)"
            ))
            
        except Exception as e:
            self.root.after(0, self.stop_processing_ui)
            self.root.after(0, lambda: self.status_label.config(
                text="✗ Error occurred during processing",
                fg="#FF4444"
            ))
            self.root.after(0, lambda: messagebox.showerror(
                "Processing Error",
                f"An error occurred during processing:\n\n{str(e)}"
            ))
            
    def save_formatted_excel(self, df, output_path):
        """Save DataFrame to Excel with HD Supply formatting"""
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Velocity Validation', index=False)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Velocity Validation']
            
            # Apply formatting
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            
            # HD Supply color scheme
            header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            header_font = Font(color="FFD700", bold=True, size=11)
            
            yellow_fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
            red_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            green_fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
            
            thin_border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
            
            # Format headers
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
                
            # Highlight Match column
            if 'Match' in df.columns:
                match_col_idx = df.columns.get_loc('Match') + 1
                for row_idx in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row_idx, column=match_col_idx)
                    cell_value = cell.value
                    
                    if cell_value == False or cell_value == 'False':
                        cell.fill = red_fill
                        cell.font = Font(color="CC0000", bold=True)
                    elif cell_value == True or cell_value == 'True':
                        cell.fill = green_fill
                        cell.font = Font(color="006600", bold=True)
                        
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border
                    
            # Highlight Current_Velocity column
            if 'Current_Velocity' in df.columns:
                vel_col_idx = df.columns.get_loc('Current_Velocity') + 1
                for row_idx in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row_idx, column=vel_col_idx)
                    cell.fill = yellow_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border
                    
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 3, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
                
    def start_processing_ui(self):
        """Update UI when processing starts"""
        self.progress.pack(pady=15)
        self.progress.start(10)
        self.process_btn.itemconfig(self.process_btn.rect, fill=self.light_gray)
        self.process_btn.unbind("<Button-1>")
        self.process_btn.configure(cursor="")
        
    def stop_processing_ui(self):
        """Update UI when processing stops"""
        self.progress.stop()
        self.progress.pack_forget()
        self.process_btn.itemconfig(self.process_btn.rect, fill=self.hd_yellow)
        self.process_btn.bind("<Button-1>", lambda e: self.process_data())
        self.process_btn.configure(cursor="hand2")
        self.status_label.config(fg=self.hd_yellow)

def main():
    root = tk.Tk()
    
    # Center window on screen
    window_width = 800
    window_height = 700
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    center_x = int(screen_width/2 - window_width/2)
    center_y = int(screen_height/2 - window_height/2)
    root.geometry(f'{window_width}x{window_height}+{center_x}+{center_y}')
    
    app = VelocityValidatorApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()
